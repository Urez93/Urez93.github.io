# -*- coding: utf-8 -*-
"""
Макрос для Компас-3D: выгружает состав изделия (спецификацию) из
активно открытой модели и сохраняет уже СТРУКТУРИРОВАННУЮ таблицу
(с отступами по уровням вложенности и группировкой строк) рядом с
открытым файлом модели — в ту же папку.

Что делает:
  1. Подключается к запущенному Компас-3D и находит активный документ.
  2. По активному документу определяет папку, где лежит файл модели —
     туда же будет сохранён результат.
  3. Находит документ спецификации: либо сам активный документ уже
     является спецификацией, либо ищет открытую спецификацию среди
     остальных открытых в Компасе документов.
  4. Экспортирует спецификацию во временный плоский excel-файл штатным
     сохранением Компаса и сразу строит из него наглядную
     структурированную таблицу через kompas_bom_structure.py.
  5. Сохраняет итоговый файл "<имя_модели>_состав_изделия.xlsx" в папку
     с открытой моделью, временный плоский файл удаляет.

Требования:
  - pywin32 (win32com) — связь с Компасом;
  - openpyxl (и xlrd, если промежуточный файл получится в старом .xls).

Файл полностью самостоятельный (логика построения структуры из
kompas_bom_structure.py включена сюда же) — это специально сделано,
потому что встроенный Python в Компасе не всегда видит соседние .py
файлы через обычный импорт (другая рабочая директория/__file__).
Для ручной пост-обработки уже сохранённых файлов вне Компаса можно
по-прежнему пользоваться отдельным kompas_bom_structure.py.

ВАЖНО про надёжность:
  Точные названия свойств/методов COM-объектов и числовые коды формата
  для экспорта спецификации в Excel могут отличаться между версиями
  Компас-3D. Поэтому для каждого шага сделано несколько попыток разными
  способами (как и в остальных макросах в этом репозитории) — в
  консоли будет видно, какой именно способ сработал, а какой нет. Если
  ни один из способов экспорта не сработает — пришлите вывод консоли,
  чтобы подобрать точный вызов для вашей версии Компаса.
"""

import os


DOCUMENT_TYPE_SPECIFICATION = 5  # ksDocumentSpecification

# ---------------------------------------------------------------------------
# Построение структурированной таблицы (та же логика, что в
# kompas_bom_structure.py, включена сюда, чтобы макрос не зависел от
# импорта соседнего файла внутри встроенного Python Компаса).
# ---------------------------------------------------------------------------

BOM_OUTPUT_COLUMNS = [
    ("Уровень", "level"),
    ("Позиция", "Позиция"),
    ("Обозначение", "Обозначение"),
    ("Наименование", "Наименование"),
    ("Количество", "Количество"),
    ("Масса", "Масса"),
    ("Материал", "Материал"),
    ("Раздел спецификации", "Раздел спецификации"),
    ("Форматы листов документа", "Форматы листов документа"),
    ("Вид изделия", "Вид изделия"),
    ("Обозначение стандарта", "Обозначение стандарта"),
    ("Типоразмер", "Типоразмер"),
    ("Примечание", "Примечание"),
]


def _read_bom_source_rows(path):
    """Читает исходную плоскую выгрузку (.xls или .xlsx) построчно."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            row = dict(zip(header, values))
            if any(str(v).strip() for v in values):
                rows.append(row)
        return rows

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb.active
    rows_iter = sh.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        row = dict(zip(header, values))
        rows.append(row)
    return rows


def _bom_clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _bom_level_from_n(n_value):
    """Уровень вложенности по коду вида '1.2.3' -> 3. Пустой/битый код -> 1."""
    text = _bom_clean(n_value)
    if not text:
        return 1
    return text.count(".") + 1


def _build_bom_tree_rows(source_rows):
    """Дополняет строки уровнем вложенности и признаком 'это узел (есть дети)'."""
    codes = [_bom_clean(r.get("N")) for r in source_rows]

    for row, code in zip(source_rows, codes):
        row["level"] = _bom_level_from_n(code)
        row["_code"] = code
        row["_is_node"] = any(
            other and other != code and other.startswith(code + ".")
            for other in codes
        )
    return source_rows


def build_structured_workbook(source_path, output_path):
    """Читает плоскую выгрузку Компаса и сохраняет наглядный xlsx с деревом."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = _read_bom_source_rows(source_path)
    rows = _build_bom_tree_rows(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Состав изделия"

    # Группировка строк сворачивается кнопкой над группой (родитель сверху).
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    node_font = Font(bold=True)
    node_fill = PatternFill("solid", fgColor="DCE6F1")

    for col_idx, (title, _) in enumerate(BOM_OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    name_col_index = [
        i for i, (_, key) in enumerate(BOM_OUTPUT_COLUMNS, start=1) if key == "Наименование"
    ][0]

    excel_row = 2
    for row in rows:
        level = row.get("level", 1)
        is_node = row.get("_is_node", False)

        for col_idx, (_, key) in enumerate(BOM_OUTPUT_COLUMNS, start=1):
            if key == "level":
                value = level
            else:
                value = _bom_clean(row.get(key, ""))
                if value == "":
                    value = None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            if col_idx == name_col_index:
                cell.alignment = Alignment(indent=min(level - 1, 14) * 2, vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

            if is_node:
                cell.font = node_font
                cell.fill = node_fill

        # Excel допускает outline_level от 0 до 7 — глубже сворачивать не даст,
        # но отступ в названии по-прежнему покажет реальный уровень.
        ws.row_dimensions[excel_row].outline_level = min(max(level - 1, 0), 7)
        excel_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(BOM_OUTPUT_COLUMNS)), excel_row - 1)

    widths = [10, 10, 20, 55, 10, 10, 30, 20, 12, 20, 22, 22, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def _log(msg):
    print(msg)


def connect_kompas():
    """Подключение к уже запущенному Компас-3D (COM)."""
    import win32com.client

    last_error = None
    for prog_id in ("Kompas.Application.7", "Kompas.Application.5", "KOMPAS.Application.5"):
        try:
            _log("Подключаемся через {}...".format(prog_id))
            app = win32com.client.Dispatch(prog_id)
            _log("  ✓ Подключено через {}".format(prog_id))
            return app
        except Exception as e:
            last_error = e
            _log("  ✗ Не удалось: {}".format(e))

    raise RuntimeError("Не удалось подключиться к Компас-3D: {}".format(last_error))


def get_active_document(app):
    """Получение активного документа."""
    try:
        doc = app.ActiveDocument
        if doc:
            return doc
    except Exception as e:
        _log("ActiveDocument недоступен: {}".format(e))

    try:
        docs = app.Documents
        if docs and docs.Count > 0:
            return docs.Item(0)
    except Exception as e:
        _log("Documents недоступен: {}".format(e))

    raise RuntimeError("Активный документ не найден. Откройте модель/сборку в Компасе.")


def get_document_path(doc):
    """Полный путь к файлу документа — несколько вариантов на случай разных версий API."""
    for attr in ("PathName", "Path", "FullFileName", "FileName"):
        try:
            value = getattr(doc, attr)
            if value:
                _log("  Путь получен через {}: {}".format(attr, value))
                return value
        except Exception:
            continue
    raise RuntimeError(
        "Не удалось получить путь к файлу активного документа. "
        "Сохраните файл на диск (он должен иметь путь) и повторите."
    )


def find_specification_document(app, active_doc):
    """Возвращает документ спецификации: сам активный документ либо один из открытых."""
    try:
        if getattr(active_doc, "DocumentType", None) == DOCUMENT_TYPE_SPECIFICATION:
            _log("Активный документ уже является спецификацией.")
            return active_doc
    except Exception as e:
        _log("Не удалось определить тип активного документа: {}".format(e))

    try:
        docs = app.Documents
        for i in range(docs.Count):
            d = docs.Item(i)
            try:
                if getattr(d, "DocumentType", None) == DOCUMENT_TYPE_SPECIFICATION:
                    _log("Найдена открытая спецификация: {}".format(getattr(d, "Name", "")))
                    return d
            except Exception:
                continue
    except Exception as e:
        _log("Не удалось перебрать открытые документы: {}".format(e))

    return None


def export_specification_to_flat_file(spec_doc, temp_path):
    """Сохраняет спецификацию во временный excel-файл штатными средствами Компаса."""
    attempts = [
        ("SaveAs(путь)", lambda: spec_doc.SaveAs(temp_path)),
        ("SaveAs(путь, 47)", lambda: spec_doc.SaveAs(temp_path, 47)),
        ("Export(путь, 47)", lambda: spec_doc.Export(temp_path, 47)),
        ("SaveAsToFormat(путь)", lambda: spec_doc.SaveAsToFormat(temp_path)),
    ]

    for name, action in attempts:
        try:
            _log("  Попытка экспорта: {}...".format(name))
            action()
            if os.path.exists(temp_path):
                _log("  ✓ Экспорт удался: {}".format(name))
                return True
        except Exception as e:
            _log("  ✗ {} не удалась: {}".format(name, e))

    return False


def run_macro():
    print()
    print("=" * 70)
    print("КОМПАС-3D — Выгрузка структурированного состава изделия")
    print("=" * 70)
    print()

    try:
        app = connect_kompas()
        active_doc = get_active_document(app)

        print()
        _log("Определяем папку активной модели...")
        model_path = get_document_path(active_doc)
        folder = os.path.dirname(model_path)
        base_name = os.path.splitext(os.path.basename(model_path))[0]

        print()
        _log("Ищем документ спецификации...")
        spec_doc = find_specification_document(app, active_doc)
        if spec_doc is None:
            raise RuntimeError(
                "Документ спецификации не найден среди открытых. "
                "Откройте спецификацию сборки в Компасе (она должна быть "
                "среди открытых документов) и запустите макрос ещё раз."
            )

        temp_flat_path = os.path.join(folder, "~{}_temp_export.xlsx".format(base_name))
        output_path = os.path.join(folder, "{}_состав_изделия.xlsx".format(base_name))

        print()
        _log("Экспортируем спецификацию во временный файл...")
        ok = export_specification_to_flat_file(spec_doc, temp_flat_path)
        if not ok:
            raise RuntimeError(
                "Не удалось экспортировать спецификацию автоматически. "
                "Сохраните её вручную через меню Компаса (Файл -> Сохранить как -> MS Excel) "
                "и запустите kompas_bom_structure.py на полученном файле."
            )

        print()
        _log("Строим структурированную таблицу...")
        build_structured_workbook(temp_flat_path, output_path)

        try:
            os.remove(temp_flat_path)
        except Exception as e:
            _log("Не удалось удалить временный файл {}: {}".format(temp_flat_path, e))

        print()
        print("=" * 70)
        print("✓ ГОТОВО")
        print("Файл сохранён: {}".format(output_path))
        print("=" * 70)
        print()

        return output_path

    except Exception as e:
        print()
        print("=" * 70)
        print("✗ ОШИБКА")
        print("=" * 70)
        print(str(e))
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    run_macro()
