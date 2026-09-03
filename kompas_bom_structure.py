# -*- coding: utf-8 -*-
"""
Преобразование выгрузки состава изделия из Компас-3D в наглядную
структурированную таблицу Excel.

Исходный экспорт Компаса кодирует вложенность узлов в один текстовый
столбец "N" (например: 1, 1.1, 1.1.1, 1.2 ...), из-за чего непонятно,
что в какой узел/подузел входит, не читая этот код построчно.

Скрипт разбирает столбец "N" и строит настоящую древовидную структуру:
  - "Наименование" сдвигается отступом на уровень вложенности;
  - узлы (сборочные единицы, у которых есть подчинённые позиции)
    выделяются жирным и заливкой;
  - строки группируются (Excel: структура/группировка по уровням),
    так что дерево можно сворачивать и разворачивать кнопками "+"/"-";
  - добавлен отдельный столбец "Уровень" с числом вложенности.

Использование как отдельного скрипта (после экспорта из Компаса):
    python kompas_bom_structure.py входной_файл.xls выходной_файл.xlsx

Использование внутри макроса Компаса — после того как исходная
плоская таблица сохранена в файл, достаточно вызвать:
    from kompas_bom_structure import build_structured_workbook
    build_structured_workbook(flat_export_path, structured_output_path)
"""

import os
import sys


# Столбцы, которые ожидаются в исходной выгрузке Компаса, и порядок,
# в котором они попадут в итоговый файл. Если каких-то столбцов в
# исходнике нет — они просто останутся пустыми, скрипт не упадёт.
OUTPUT_COLUMNS = [
    ("Уровень", "level"),
    ("Позиция", "Позиция"),
    ("Обозначение", "Обозначение"),
    ("Наименование", "Наименование"),
    ("Количество", "Количество"),
    ("Масса", "Масса"),
    ("Материал", "Материал"),
    ("Раздел спецификации", "Раздел спецификации"),
    ("Форматы листов документа", "Форматы листов документа"),
    ("Вид изделия", "Вид изделия"),
    ("Обозначение стандарта", "Обозначение стандарта"),
    ("Типоразмер", "Типоразмер"),
    ("Примечание", "Примечание"),
]


def read_source_rows(path):
    """Читает исходную плоскую выгрузку (.xls или .xlsx) построчно."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            values = [sh.cell_value(r, c) for c in range(sh.ncols)]
            row = dict(zip(header, values))
            if any(str(v).strip() for v in values):
                rows.append(row)
        return rows

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb.active
    rows_iter = sh.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        row = dict(zip(header, values))
        rows.append(row)
    return rows


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _level_from_n(n_value):
    """Уровень вложенности по коду вида '1.2.3' -> 3. Пустой/битый код -> 1."""
    text = _clean(n_value)
    if not text:
        return 1
    return text.count(".") + 1


def build_tree_rows(source_rows):
    """Дополняет строки уровнем вложенности и признаком 'это узел (есть дети)'."""
    codes = [_clean(r.get("N")) for r in source_rows]

    for row, code in zip(source_rows, codes):
        row["level"] = _level_from_n(code)
        row["_code"] = code
        row["_is_node"] = any(
            other and other != code and other.startswith(code + ".")
            for other in codes
        )
    return source_rows


def build_structured_workbook(source_path, output_path):
    """Читает плоскую выгрузку Компаса и сохраняет наглядный xlsx с деревом."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = read_source_rows(source_path)
    rows = build_tree_rows(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Состав изделия"

    # Группировка строк сворачивается кнопкой над группой (родитель сверху).
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    node_font = Font(bold=True)
    node_fill = PatternFill("solid", fgColor="DCE6F1")

    for col_idx, (title, _) in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    name_col_index = [i for i, (_, key) in enumerate(OUTPUT_COLUMNS, start=1) if key == "Наименование"][0]

    excel_row = 2
    for row in rows:
        level = row.get("level", 1)
        is_node = row.get("_is_node", False)

        for col_idx, (_, key) in enumerate(OUTPUT_COLUMNS, start=1):
            if key == "level":
                value = level
            else:
                value = _clean(row.get(key, ""))
                if value == "":
                    value = None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            if col_idx == name_col_index:
                cell.alignment = Alignment(indent=min(level - 1, 14) * 2, vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

            if is_node:
                cell.font = node_font
                cell.fill = node_fill

        # Excel допускает outline_level от 0 до 7 — глубже сворачивать не даст,
        # но отступ в названии по-прежнему покажет реальный уровень.
        ws.row_dimensions[excel_row].outline_level = min(max(level - 1, 0), 7)
        excel_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(OUTPUT_COLUMNS)), excel_row - 1)

    widths = [10, 10, 20, 55, 10, 10, 30, 20, 12, 20, 22, 22, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) != 3:
        print("Использование: python kompas_bom_structure.py входной_файл.xls выходной_файл.xlsx")
        return 1

    source_path, output_path = sys.argv[1], sys.argv[2]
    build_structured_workbook(source_path, output_path)
    print("Готово: {}".format(output_path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
